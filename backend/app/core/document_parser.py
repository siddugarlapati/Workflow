"""
WorkFlow — Document Parser Utility
==================================
Extracts structured text from uploaded PDF, Word (.docx), and Excel (.xlsx) files
to feed them directly to the LLM Verification Engine.
"""
import io
import structlog
from typing import Optional

logger = structlog.get_logger()


def extract_text_from_pdf(file_bytes: bytes) -> str:
    """Extracts text page-by-page from a PDF byte stream."""
    try:
        from pypdf import PdfReader
        import io
        
        pdf_file = io.BytesIO(file_bytes)
        reader = PdfReader(pdf_file)
        pages_text = []
        
        for idx, page in enumerate(reader.pages):
            text = page.extract_text()
            if text:
                pages_text.append(f"--- Page {idx+1} ---\n{text}")
                
        extracted = "\n".join(pages_text)
        if not extracted.strip() and len(reader.pages) > 0:
            # Blank Scans Loophole: It's a scanned vectorless PDF!
            import base64
            b64_pdf = base64.b64encode(file_bytes).decode("utf-8")
            logger.info("parser.scanned_pdf_intercepted", pages=len(reader.pages))
            return (
                f"[SCANNED VECTORLESS PDF DETECTED (Blank Scan Loophole Intercepted)]\n"
                f"[PDF PAGES: {len(reader.pages)}]\n"
                f"[BASE64 PDF DOCUMENT DATA START]\n{b64_pdf}\n[BASE64 PDF DOCUMENT DATA END]\n"
                f"[SELF-CORRECTIVE ANALYSIS]: The PDF text stream was blank. Intercepted scanned vectorless loophole. Encompasses Base64 payload for Vision LLM auditing."
            )
        return extracted
    except Exception as e:
        logger.error("parser.pdf_failed", error=str(e))
        raise ValueError(f"Failed to parse PDF document: {e}")


def extract_text_from_docx(file_bytes: bytes) -> str:
    """Extracts text paragraph-by-paragraph from a Word (.docx) byte stream."""
    try:
        from docx import Document
        
        docx_file = io.BytesIO(file_bytes)
        doc = Document(docx_file)
        paragraphs_text = []
        
        for p in doc.paragraphs:
            if p.text.strip():
                paragraphs_text.append(p.text)
                
        return "\n".join(paragraphs_text)
    except Exception as e:
        logger.error("parser.docx_failed", error=str(e))
        raise ValueError(f"Failed to parse Word document: {e}")


def extract_text_from_xlsx(file_bytes: bytes) -> str:
    """Extracts cells row-by-row from all sheets in an Excel (.xlsx) byte stream."""
    try:
        import openpyxl
        
        xlsx_file = io.BytesIO(file_bytes)
        wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        sheets_text = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_rows = []
            
            for row in sheet.iter_rows(values_only=True):
                # Filter out empty rows
                row_values = [str(val) if val is not None else "" for val in row]
                if any(row_values):
                    sheet_rows.append(" | ".join(row_values))
                    
            if sheet_rows:
                sheets_text.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(sheet_rows))
                
        return "\n".join(sheets_text)
    except Exception as e:
        logger.error("parser.xlsx_failed", error=str(e))
        raise ValueError(f"Failed to parse Excel spreadsheet: {e}")


def extract_text_from_image(file_bytes: bytes, filename: str) -> str:
    """
    Extracts text from uploaded images (screenshots, logs).
    Routes to EasyOCR or PyTesseract dynamically if installed.
    Otherwise, extracts structural metadata and base64-encodes it for Vision LLMs.
    """
    logger.info("parser.image_detected", filename=filename, size=len(file_bytes))
    
    # 1. Try EasyOCR
    try:
        import easyocr
        reader = easyocr.Reader(['en'])
        results = reader.readtext(file_bytes, detail=0)
        if results:
            return "\n".join(results)
    except ImportError:
        pass
    except Exception as e:
        logger.warning("parser.easyocr_failed", error=str(e))
        
    # 2. Try PyTesseract
    try:
        import io
        import pytesseract
        from PIL import Image
        img = Image.open(io.BytesIO(file_bytes))
        extracted = pytesseract.image_to_string(img)
        if extracted.strip():
            return extracted
    except ImportError:
        pass
    except Exception as e:
        logger.warning("parser.tesseract_failed", error=str(e))
        
    # 3. Fallback: Base64 encode for Vision LLMs
    import base64
    b64_str = base64.b64encode(file_bytes).decode("utf-8")
    
    return (
        f"[IMAGE SCREENSHOT PROOF UPLOADED: {filename}]\n"
        f"[IMAGE SIZE: {len(file_bytes)} bytes]\n"
        f"[BASE64 IMAGE ENCODED STREAM DATA START]\n{b64_str}\n[BASE64 IMAGE ENCODED STREAM DATA END]\n"
        f"[SELF-CORRECTIVE ANALYSIS]: This image represents visual proof (screenshot). It has been base64-encoded for LLM vision-analysis matching."
    )


def extract_text_from_document(file_bytes: bytes, filename: str) -> str:
    """
    Gateway function: Routes file to the appropriate parser based on extension.
    Falls back to decoding as UTF-8 string for plaintext/CSV formats.
    """
    ext = filename.split(".")[-1].lower()
    
    if ext == "pdf":
        return extract_text_from_pdf(file_bytes)
    elif ext in ("doc", "docx"):
        return extract_text_from_docx(file_bytes)
    elif ext in ("xls", "xlsx"):
        return extract_text_from_xlsx(file_bytes)
    elif ext in ("png", "jpg", "jpeg", "webp"):
        return extract_text_from_image(file_bytes, filename)
    elif ext in ("txt", "csv", "json"):
        try:
            return file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            try:
                return file_bytes.decode("latin-1")
            except Exception:
                raise ValueError("Plaintext document contains unsupported encoding.")
    else:
        raise ValueError(f"Unsupported file format: '.{ext}'. Supported: PDF, DOCX, XLSX, TXT, CSV, PNG, JPG, JPEG, WEBP")
